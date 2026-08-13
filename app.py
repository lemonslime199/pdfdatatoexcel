import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="PDF Statement Extractor", page_icon="💳", layout="wide")

st.title("💳 ระบบดึงข้อมูลรายการใช้บัตรเครดิตลง Excel (Professional Edition)")
st.write("ดึงคอลัมน์: **วันที่ใช้บัตร**, **วันที่บันทึกรายการ**, **รายการ**, และ **จำนวนเงิน** (รองรับยอดติดลบ/เครดิตคืนเงิน แยก Sheet ตาม **วันที่ครบกำหนดชำระ**)")

uploaded_files = st.file_uploader(
    "เลือกหรือลากไฟล์ PDF ใบแจ้งยอดมาวางที่นี่ (รองรับหลายไฟล์พร้อมกัน)", 
    type=["pdf"], 
    accept_multiple_files=True
)

if uploaded_files:
    st.write(f"📁 จำนวนไฟล์ที่อัปโหลด: **{len(uploaded_files)}** ไฟล์")
    
    file_data_map = {}
    
    with st.spinner("กำลังประมวลผลและจัดรูปแบบตารางอย่างสวยงาม..."):
        for uploaded_file in uploaded_files:
            recording = False
            extracted_records = []
            due_date = "Sheet_Data"
            
            with pdfplumber.open(uploaded_file) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    
                    lines = text.split("\n")
                    
                    for line in lines:
                        clean_line = line.strip()
                        
                        # ค้นหาวันที่ครบกำหนดชำระ
                        if ("ครบกำหนด" in clean_line or "Due Date" in clean_line) and due_date == "Sheet_Data":
                            date_match = re.search(r'(\d{2}[/.-]\d{2}[/.-]\d{2,4})', clean_line)
                            if date_match:
                                raw_date = date_match.group(1).replace('/', '-').replace('.', '-')
                                due_date = f"ครบชำระ_{raw_date}"
                        
                        # เริ่มอ่านเมื่อพบข้อความ "วันที่ใช้บัตร"
                        if "วันที่ใช้บัตร" in clean_line:
                            recording = True
                            continue
                        
                        if recording and clean_line:
                            parts = clean_line.split()
                            
                            # เจอคำว่า "สรุปยอด" -> เก็บบรรทัดสรุปยอดแล้วจบการอ่าน
                            if "สรุปยอด" in clean_line:
                                if len(parts) >= 2:
                                    summary_label = " ".join([p for p in parts if not p.replace(',', '').replace('.', '').replace('-', '').replace('CR', '').replace('CR.', '').isdigit()])
                                    summary_amount = parts[-1]
                                    
                                    extracted_records.append({
                                        "วันที่ใช้บัตร": "-",
                                        "วันที่บันทึกรายการ": "-",
                                        "รายการ": summary_label if summary_label else "สรุปยอดงวดนี้",
                                        "จำนวนเงิน": summary_amount
                                    })
                                recording = False
                                break
                            
                            # เก็บบรรทัดรายการปกติ
                            if len(parts) >= 4:
                                txn_date = parts[0]
                                post_date = parts[1]
                                amount = parts[-1]
                                description = " ".join(parts[2:-1])
                                
                                extracted_records.append({
                                    "วันที่ใช้บัตร": txn_date,
                                    "วันที่บันทึกรายการ": post_date,
                                    "รายการ": description,
                                    "จำนวนเงิน": amount
                                })
            
            if extracted_records:
                sheet_name = due_date[:31]
                base_name = sheet_name
                counter = 1
                while sheet_name in file_data_map:
                    sheet_name = f"{base_name}_{counter}"[:31]
                    counter += 1
                
                file_data_map[sheet_name] = pd.DataFrame(extracted_records)

    if file_data_map:
        st.success("✅ ดึงและจัดกลุ่มข้อมูลสำเร็จ!")
        
        # แสดงตัวอย่างข้อมูลตามแท็บหน้าเว็บ Streamlit
        tabs = st.tabs(list(file_data_map.keys()))
        for tab, (s_name, df_data) in zip(tabs, file_data_map.items()):
            with tab:
                st.write(f"### 📑 รายการสำหรับแท็บ: **{s_name}**")
                st.dataframe(df_data, use_container_width=True)
        
        # สร้างไฟล์ Excel
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            for s_name, df_data in file_data_map.items():
                # เขียนข้อมูลลง Sheet เริ่มแถวที่ 5
                df_data.to_excel(writer, index=False, sheet_name=s_name, startrow=4)
                
                worksheet = writer.sheets[s_name]
                worksheet.views.sheetView[0].showGridLines = True
                
                # Title Block (พรีเมียม สไตล์ Modern Corporate)
                worksheet.merge_cells("A1:D1")
                title_cell = worksheet["A1"]
                title_cell.value = "💳 รายงานสรุปรายการใช้บัตรเครดิต"
                title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="0F2C59")
                title_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                worksheet.merge_cells("A2:D2")
                sub_cell = worksheet["A2"]
                sub_cell.value = f"วันที่ครบกำหนดชำระ: {s_name.replace('ครบชำระ_', '')}  |  ดึงข้อมูลอัตโนมัติจาก PDF"
                sub_cell.font = Font(name="Segoe UI", size=10, italic=True, color="4F709C")
                sub_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Header Styling (Row 5 - Deep Slate Blue)
                header_fill = PatternFill(start_color="0F2C59", end_color="0F2C59", fill_type="solid")
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
                
                worksheet.row_dimensions[5].height = 28
                headers = ["วันที่ใช้บัตร", "วันที่บันทึกรายการ", "รายการ", "จำนวนเงิน (บาท)"]
                for col_idx, text in enumerate(headers, start=1):
                    cell = worksheet.cell(row=5, column=col_idx, value=text)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin_border
                
                # Palette สี
                zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                total_row_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                
                neg_font = Font(name="Segoe UI", size=11, bold=True, color="C0392B")
                pos_font = Font(name="Segoe UI", size=11, color="1A252C")
                total_font = Font(name="Segoe UI", size=11, bold=True, color="0F2C59")
                
                max_r = 5 + len(df_data)
                for r_idx in range(6, max_r + 1):
                    worksheet.row_dimensions[r_idx].height = 22
                    is_last_row = (r_idx == max_r)
                    
                    # Alignment
                    worksheet.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    worksheet.cell(row=r_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
                    worksheet.cell(row=r_idx, column=3).alignment = Alignment(horizontal="left", vertical="center")
                    
                    amt_cell = worksheet.cell(row=r_idx, column=4)
                    amt_cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # แปลงจำนวนเงิน (รองรับค่าลบ / CR)
                    is_negative = False
                    try:
                        raw_str = str(amt_cell.value).strip()
                        if 'CR' in raw_str.upper() or '-' in raw_str:
                            is_negative = True
                        
                        clean_val = raw_str.replace(',', '').replace('฿', '').replace('CR', '').replace('cr', '').replace('-', '').strip()
                        num_val = float(clean_val)
                        if is_negative:
                            num_val = -abs(num_val)
                        
                        amt_cell.value = num_val
                        amt_cell.number_format = '#,##0.00;[Red]-#,##0.00;0.00'
                    except:
                        pass
                    
                    # ตกแต่งสีและขอบ
                    for c_idx in range(1, 5):
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.border = thin_border
                        
                        if is_last_row:
                            cell.fill = total_row_fill
                            cell.font = total_font
                        else:
                            cell.font = pos_font
                            if r_idx % 2 == 1:
                                cell.fill = zebra_fill
                    
                    # เน้นตัวหนังสือสีแดงสำหรับยอดติดลบ
                    if is_negative and not is_last_row:
                        amt_cell.font = neg_font

                # ความกว้างคอลัมน์
                worksheet.column_dimensions['A'].width = 18
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 48
                worksheet.column_dimensions['D'].width = 22

        st.download_button(
            label="📥 ดาวน์โหลดไฟล์ Excel สวยงามระดับมืออาชีพ (.xlsx)",
            data=excel_buffer.getvalue(),
            file_name="statement_formatted_pro.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("⚠️ ไม่พบข้อมูลรายการใช้บัตรในไฟล์ PDF ที่อัปโหลด กรุณาตรวจสอบไฟล์อีกครั้ง")
